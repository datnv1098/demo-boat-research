#!/usr/bin/env python3
"""
Convert Excel data from ตัวอย่างข้อมูลทร้พยากรสัตว์น้ำจากเรือสำ.xlsx
into the JSON format expected by converted_excel_data.json

Target structure:
  {
    "trips": [...],
    "cpueData": [...],
    "lengthData": [...],
    "waterQualityData": [...],
    "speciesInfo": {...}
  }
"""
import json
import openpyxl
from datetime import datetime, date

EXCEL_PATH = "document/ตัวอย่างข้อมูลทร้พยากรสัตว์น้ำจากเรือสำ.xlsx"
OUTPUT_PATH = "converted_excel_data.json"

def load_workbook():
    return openpyxl.load_workbook(EXCEL_PATH, data_only=True)

def read_sheet_as_dicts(wb, sheet_name):
    """Read a sheet into list of dicts using row 1 as headers."""
    ws = wb[sheet_name]
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        h = str(cell).strip().replace('\ufeff', '') if cell else f'col_{len(headers)}'
        headers.append(h)
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        d = {}
        for i, val in enumerate(row):
            if i < len(headers):
                d[headers[i]] = val
        rows.append(d)
    return rows

def format_date(val):
    """Convert datetime/date/string to YYYY-MM-DD string."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    if val is None:
        return ""
    s = str(val).strip()
    # Handle M-DD-YYYY or MM-DD-YYYY format
    import re
    m = re.match(r'^(\d{1,2})-(\d{1,2})-(\d{4})$', s)
    if m:
        month, day, year = m.groups()
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    # Handle MM/DD/YYYY
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        month, day, year = m.groups()
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    return s

def map_fishing_area(main_area, centre_id):
    """Map Main_Area + centre_id to Thai fishing area description."""
    if main_area and str(main_area).upper() in ('AND', 'ADM'):
        if centre_id == 'AFD':
            return "ฝั่งอันดามันใต้ (ภูเก็ต-สตูล)"
        return "ฝั่งอันดามันเหนือ (ระนอง-พังงา)"
    # GOT
    if centre_id == 'EMD':
        return "อ่าวไทยตอนบน (ระยอง-ตราด)"
    if centre_id == 'UMD':
        return "อ่าวไทยตอนบน (สมุทรปราการ)"
    if centre_id == 'CMD':
        return "อ่าวไทยตอนกลาง (ชุมพร-ประจวบฯ)"
    if centre_id == 'SMD':
        return "อ่าวไทยตอนล่าง (สงขลา-ปัตตานี)"
    return "อ่าวไทย"

def map_habitat(group_species):
    """Map Group_species to habitat description."""
    if not group_species:
        return ""
    gs = str(group_species).lower()
    if 'pelagic' in gs:
        return "ผิวน้ำ, ฝูงปลา"
    if 'demersal' in gs:
        return "หน้าดิน"
    if 'cephalopod' in gs or 'squid' in gs:
        return "น้ำลึก, ผิวน้ำ"
    if 'shrimp' in gs or 'prawn' in gs:
        return "หน้าดิน, ชายฝั่ง"
    if 'crab' in gs:
        return "หน้าดิน, ชายฝั่ง"
    if 'trash' in gs or 'juvenile' in gs:
        return "หน้าดิน"
    return "ทะเล"

def map_economic_value(fish_group, group_species):
    """Map fish_group/Group_species to economic value."""
    if not group_species:
        return "ปานกลาง"
    gs = str(group_species).lower()
    if 'trash' in gs or 'juvenile' in gs:
        return "ต่ำ"
    if fish_group and int(fish_group) <= 2 if str(fish_group).isdigit() else False:
        return "สูง"
    if 'pelagic' in gs or 'cephalopod' in gs:
        return "สูง"
    if 'shrimp' in gs or 'prawn' in gs:
        return "สูง"
    return "ปานกลาง"

def main():
    print("Loading workbook...")
    wb = load_workbook()

    # ── 1. Load reference tables ──
    print("Reading sheets...")
    centers = read_sheet_as_dicts(wb, 'center')
    areas = read_sheet_as_dicts(wb, 'area')
    efforts = read_sheet_as_dicts(wb, 'effort2')
    catches = read_sheet_as_dicts(wb, 'catch2')
    species_list = read_sheet_as_dicts(wb, 'species')

    # Build species lookup by dpSP_id
    species_by_id = {}
    for sp in species_list:
        dpid = sp.get('dpSP_id')
        if dpid is not None:
            species_by_id[int(dpid)] = sp

    # Build center lookup
    center_by_id = {}
    for c in centers:
        center_by_id[c.get('center_id', '')] = c

    # Build area lookup by Station (Stat_area)
    area_by_station = {}
    for a in areas:
        rv_st = a.get('RV_St_ID')
        if rv_st is not None:
            area_by_station[int(rv_st)] = a

    # Build effort lookup by sample_id
    effort_by_sample = {}
    for e in efforts:
        sid = e.get('sample_id')
        if sid:
            effort_by_sample[str(sid)] = e

    # ── 2. Pre-compute totalCatch per sample_id from catch2 ──
    print("Computing total catches per trip...")
    total_catch_by_sample = {}
    for c in catches:
        sid = str(c.get('sample_id', ''))
        tw = c.get('total_weight')
        if tw and sid:
            try:
                w = float(tw)
                if w > 0:
                    total_catch_by_sample[sid] = total_catch_by_sample.get(sid, 0) + w
            except (ValueError, TypeError):
                pass

    # ── 3. Build trips[] from effort2 ──
    print("Building trips...")
    trips = []
    for e in efforts:
        sample_id = str(e.get('sample_id', ''))
        if not sample_id:
            continue
        centre_id = str(e.get('centre_id', ''))
        main_area = str(e.get('Main_Area', ''))
        vessel_name = str(e.get('VesselName', ''))
        sample_date = e.get('Sample_Date_Eng')
        date_str = format_date(sample_date)
        lat = e.get('LatStart')
        lon = e.get('LongStart')
        tow_time = e.get('Tow_Time')
        station = e.get('Station')
        depth = e.get('Depth')
        rv_year = e.get('RV_Year')
        rv_month = e.get('RV_Month')
        haulno = e.get('Haulno')

        fishing_area = map_fishing_area(main_area, centre_id)
        total_catch = round(total_catch_by_sample.get(sample_id, 0), 2)
        center_info = center_by_id.get(centre_id, {})
        remark = e.get('Remark')

        # Use DQ score and issues from the Excel source directly.
        # The expert department has already validated this data,
        # so we trust it as-is (dqScore=100, issues=[]) unless
        # the Remark field contains annotations.
        dq = 100
        issues = []
        if remark:
            # Remark from expert: lower score and surface it
            dq = 80
            issues.append(str(remark))

        trips.append({
            "tripId": sample_id,
            "vessel": vessel_name,
            "vesselType": "เรือสำรวจประมง",
            "captain": "",
            "startDate": date_str,
            "endDate": date_str,
            "fishingArea": fishing_area,
            "coordinates": {
                "lat": float(lat) if lat else 0,
                "lon": float(lon) if lon else 0
            },
            "duration": int(tow_time) if tow_time else 0,
            "dqScore": dq,
            "issues": issues,
            "totalCatch": total_catch,
            "fuelConsumption": 0,
            "station": int(station) if station else 0,
            "depth": float(depth) if depth else 0,
            "cruise": int(e.get('Cruise', 0)) if e.get('Cruise') else 0,
            "haulNo": int(haulno) if haulno else 0,
            "rvYear": int(rv_year) if rv_year else 0,
            "rvMonth": int(rv_month) if rv_month else 0,
            "centerId": centre_id,
            "centerName": center_info.get('centerName', '')
        })

    # ── 4. Build cpueData[] from catch2 + effort2 ──
    print("Building CPUE data...")
    cpue_data = []
    for c in catches:
        sid = str(c.get('sample_id', ''))
        species_id = c.get('species_id')
        tw = c.get('total_weight')

        if not sid or tw is None:
            continue
        try:
            weight = float(tw)
        except (ValueError, TypeError):
            continue
        if weight <= 0:
            continue

        # Look up effort info
        effort = effort_by_sample.get(sid, {})
        sample_date = effort.get('Sample_Date_Eng')
        date_str = format_date(sample_date)
        if not date_str:
            continue

        month_str = date_str[:7] if len(date_str) >= 7 else ""
        centre_id = str(effort.get('centre_id', ''))
        main_area = str(effort.get('Main_Area', ''))
        fishing_area = map_fishing_area(main_area, centre_id)
        tow_time = effort.get('Tow_Time', 60)

        # Look up species Thai name
        sp_info = species_by_id.get(int(species_id), {}) if species_id else {}
        th_name = sp_info.get('Th_name', '')
        sci_name = sp_info.get('Scientific_name', '')
        species_name = th_name if th_name else sci_name if sci_name else f"species_{species_id}"

        # CPUE = catch / effort (standardized per hour)
        effort_hours = float(tow_time) / 60.0 if tow_time else 1.0
        cpue_val = round(weight / effort_hours, 3) if effort_hours > 0 else 0

        cpue_data.append({
            "tripId": sid,
            "date": date_str,
            "month": month_str,
            "fishingArea": fishing_area,
            "species": species_name,
            "cpue": cpue_val,
            "effort": effort_hours,
            "catch": round(weight, 3),
            "station": int(effort.get('Station', 0)) if effort.get('Station') else 0
        })

    # ── 5. Build lengthData[] from catch2 sex/number data ──
    # The Excel doesn't have explicit length-frequency bins,
    # but we can build records from the sex and number fields in catch2
    print("Building length data from catch records...")
    length_data = []
    for c in catches:
        species_id = c.get('species_id')
        sex = c.get('sex', 'NA')
        number = c.get('Number')
        sample_weight = c.get('sample_weight')
        cod_cover = c.get('cod_cover', '')
        sid = str(c.get('sample_id', ''))
        rv_month = c.get('RV_Month')

        if number is None or species_id is None:
            continue
        try:
            num = int(number)
        except (ValueError, TypeError):
            continue
        if num <= 0:
            continue

        sp_info = species_by_id.get(int(species_id), {}) if species_id else {}
        th_name = sp_info.get('Th_name', '')
        sci_name = sp_info.get('Scientific_name', '')
        species_name = th_name if th_name else sci_name if sci_name else f"species_{species_id}"

        effort = effort_by_sample.get(sid, {})
        main_area = str(effort.get('Main_Area', ''))
        centre_id = str(effort.get('centre_id', ''))

        # Map area
        if main_area.upper() in ('AND', 'ADM'):
            area_label = "ฝั่งอันดามัน"
        else:
            area_label = "อ่าวไทย"

        # Season from month
        month = rv_month if rv_month else effort.get('RV_Month', 1)
        if month:
            m = int(month)
            if m <= 3:
                season = "Q1"
            elif m <= 6:
                season = "Q2"
            elif m <= 9:
                season = "Q3"
            else:
                season = "Q4"
        else:
            season = "Q1"

        # Estimate length bin from sample_weight and number
        if sample_weight and num > 0:
            try:
                avg_weight_g = float(sample_weight) * 1000 / num
                # Rough length estimate (cm) from weight using L ~ W^(1/3) * factor
                est_length = round((avg_weight_g ** 0.33) * 1.5, 0)
                bin_low = max(1, int(est_length - 1))
                bin_high = int(est_length + 1)
                length_bin = f"{bin_low}-{bin_high}ซม."
            except (ValueError, TypeError):
                length_bin = "ไม่ระบุ"
        else:
            length_bin = "ไม่ระบุ"

        male = num if sex == 'M' else 0
        female = num if sex == 'F' else 0
        unsexed = num if sex not in ('M', 'F') else 0

        length_data.append({
            "species": species_name,
            "lengthBin": length_bin,
            "male": male,
            "female": female,
            "unsexed": unsexed,
            "area": area_label,
            "season": season,
            "codCover": str(cod_cover) if cod_cover else ""
        })

    # ── 6. Build waterQualityData[] ──
    # No water quality data in the Excel; keep empty array
    water_quality_data = []

    # ── 7. Build speciesInfo{} from species sheet ──
    print("Building species info...")
    species_info = {}
    seen_ids = set()
    for sp in species_list:
        th_name = sp.get('Th_name', '')
        sci_name = sp.get('Scientific_name', '')
        common_name = sp.get('Common_name', '')
        group_species = sp.get('Group_species', '')
        fish_group = sp.get('fish_group', '')
        dpid = sp.get('dpSP_id')

        if not th_name and not sci_name:
            continue
        if dpid in seen_ids:
            continue
        seen_ids.add(dpid)

        key = th_name if th_name else sci_name
        if key in species_info:
            continue

        species_info[key] = {
            "scientificName": sci_name if sci_name else "",
            "commonName": common_name if common_name else "",
            "thaiName": th_name if th_name else "",
            "family": sp.get('Family_name', '') or "",
            "genus": sp.get('Genus', '') or "",
            "groupSpecies": str(group_species) if group_species else "",
            "statNameTh": sp.get('Stat_name_th', '') or "",
            "statNameEng": sp.get('Stat_name_eng', '') or "",
            "lm50": 0,
            "maxLength": 0,
            "habitat": map_habitat(group_species),
            "economicValue": map_economic_value(fish_group, group_species)
        }

    # ── 8. Assemble and write ──
    result = {
        "trips": trips,
        "cpueData": cpue_data,
        "lengthData": length_data,
        "waterQualityData": water_quality_data,
        "speciesInfo": species_info
    }

    print(f"Writing JSON: {len(trips)} trips, {len(cpue_data)} cpueData, "
          f"{len(length_data)} lengthData, {len(water_quality_data)} waterQualityData, "
          f"{len(species_info)} species")

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"Done! Output written to {OUTPUT_PATH}")

if __name__ == '__main__':
    main()
